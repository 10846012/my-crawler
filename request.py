from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl


def launchBrowser():
    PATH = Service(r"C:\Users\Yun\Desktop\me\crawler\chromedriver.exe")

    chrome = webdriver.Chrome(service=PATH)
    chrome.get("https://www.instagram.com/")
    time.sleep(2)
    username = chrome.find_element(
        By.XPATH, '//input[@name="username"]')
    username.send_keys('')
    password = chrome.find_element(
        By.XPATH, '//input[@name="password"]')
    password.send_keys('')
    login_click = chrome.find_element(
        By.XPATH, '//*[@id="loginForm"]/div/div[3]')
    login_click.click()

    time.sleep(5)
    chrome.get("https://www.instagram.com/nba/")
    time.sleep(7)
    data = []

    actions = ActionChains(chrome)
    for i in range(5):
        post = chrome.find_elements(
            By.XPATH, '//div[@class="_aabd _aa8k  _al3l"]/a')
        href = post[i].get_attribute('href')
        actions.move_to_element(post[i]).perform()
        time.sleep(3)
        likesAndComments = chrome.find_elements(
            By.XPATH, '//li[@class="_abpm"]/span/span')
        likes = likesAndComments[0].text
        likes = likes.encode("utf8").decode("cp950", "ignore")
        comments = likesAndComments[1].text
        comments = comments.encode("utf8").decode("cp950", "ignore")

        dic = {'href': href, 'likes': likes, 'commentsNum': comments}
        data.append(dic)
        if (len(data) >= 5):
            break
    commentList = []
    for k in range(5):

        chrome.get(data[k]['href'])
        time.sleep(5)
        postTime = chrome.find_element(
            By.XPATH, '//time[@class="_aaqe"]')
        postTime = postTime.get_attribute('datetime')
        data[k]['time'] = postTime.encode("utf8").decode("cp950", "ignore")
        content = chrome.find_elements(
            By.XPATH, '//h1[@class="_aacl _aaco _aacu _aacx _aad7 _aade"]')[0].text
        data[k]['content'] = content.encode("utf8").decode("cp950", "ignore")
        conment = chrome.find_elements(
            By.XPATH, '//span[@class="_aacl _aaco _aacu _aacx _aad7 _aade"]')
        conment_len = len(conment)
        for i in range(conment_len-1):
            conment = chrome.find_elements(
                By.XPATH, '//span[@class="_aacl _aaco _aacu _aacx _aad7 _aade"]')[i].text
            commentList.append(conment.encode(
                "utf8").decode("cp950", "ignore"))
        data[k]['comments'] = commentList
        commentList = []

    export(data)




def export(data):
    wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
    wb.save('crawler_excel.xlsx') 

    wb = openpyxl.load_workbook('crawler_excel.xlsx', data_only=True) 

    s1 = wb['Sheet']

    format =[['','貼文連結','讚數(萬)','留言數','發文時間','文章標題','留言'],['1','','','','','',''],['2','','','','','',''],['3','','','','','',''],['4','','','','','',''],['5','','','','','','']]
    for i in format:
        s1.append(i)

    data2 = data

    for y in range(len(data2)):
        row = 2 + y
        s1.cell(row,1).value = data2[y]['href']
        s1.cell(row,2).value = data2[y]['likes']
        s1.cell(row,3).value = data2[y]['commentsNum']
        s1.cell(row,4).value = data2[y]['time']
        s1.cell(row,5).value = data2[y]['content'].encode(
                "utf8").decode("cp950", "ignore")
        commentlen = len(data2[y]['comments'])
        for j in range(commentlen):
            col = 5+j
            s1.cell(row,col).value = data2[y]['comments'][j].encode(
                    "utf8").decode("cp950", "ignore")
    wb.save('crawler_excel.xlsx')


switch = True

driver = launchBrowser()